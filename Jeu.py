import random
import os
import openpyxl as px
import time 
from P4 import * 
from MCTS import *
from Minmax import *

MCTS_ITERATION = 4000


def returnString(grille):
    ret = ""
    for i in range(6):
        for j in range(7):
            ret += "[" + grille[j][5 - i] + "]"
        ret += "\n"
    ret += " ========================================"
    ret += "\n"
    return ret


def jouerMinMax(grille, max_depth):
    joue = True
    no_tour = 0
    if(MINMAX_COULEUR==couleur.jaune):
        while joue:
            print(no_tour)
            if (no_tour % 2) == 0:
                coup = humain_joue(grille)
                grille = jouer_coup(couleur.rouge,coup,grille)
                joue = not (verif_gagnage(couleur.rouge, grille))
            else:
                coup = robot_joue_mieux(liste_coup_possible(grille), grille, max_depth)
                grille = jouer_coup(couleur.jaune, coup, grille)
                joue = not (verif_gagnage(couleur.jaune, grille))
            afficher_grille(grille)
            no_tour = no_tour + 1
    else:
        while joue:
            print(no_tour)
            if (no_tour % 2) != 0:
                coup = humain_joue(grille)
                grille = jouer_coup(couleur.jaune,coup,grille)
                joue = not (verif_gagnage(couleur.jaune, grille))
            else:
                coup = robot_joue_mieux(liste_coup_possible(grille), grille, max_depth)
                grille = jouer_coup(couleur.rouge, coup, grille)
                joue = not (verif_gagnage(couleur.rouge, grille))
            afficher_grille(grille)
            no_tour = no_tour + 1


def jouerMCTS(grille, iteration) : 
    joue = True
    no_tour = 0
    noeud_racine = Node(grille, None, 0, None)
    if(MCTS_COULEUR==couleur.jaune):
        while joue :
            if (no_tour % 2) == 0 :           
                coup = humain_joue(grille)
                grille = jouer_coup(couleur.rouge, coup, grille)
                joue = not (verif_gagnage(couleur.rouge, grille))
                if (len(noeud_racine.enfants)>0) : 
                    no_enfant = noeud_racine.retrouveEnfant(coup)
                    noeud_racine = noeud_racine.enfants[no_enfant]
                    noeud_racine.devient_racine()
                
                else : 
                    noeud_racine = Node(grille, None, 1, None)
            else :
                coup = robot_joue_encore_mieux(noeud_racine, MCTS_COULEUR, iteration)[0]
                print (coup, " coup joué")
                afficher_grille(noeud_racine.etat)
                grille = jouer_coup(couleur.jaune, coup, grille)
                joue = not (verif_gagnage(couleur.jaune, grille))
                if (len(noeud_racine.enfants)>0) : 
                    no_enfant = noeud_racine.retrouveEnfant(coup)
                    noeud_racine = noeud_racine.enfants[no_enfant]
                    noeud_racine.devient_racine()               
                else : 
                    noeud_racine = Node(grille, None, 0, None)
            afficher_grille(grille)
            no_tour = no_tour + 1
    else:
        while joue :
            if (no_tour % 2) != 0 :           
                coup = humain_joue(grille)
                grille = jouer_coup(couleur.jaune, coup, grille)
                joue = not (verif_gagnage(couleur.jaune, grille))
                if (len(noeud_racine.enfants)>0) : 
                    no_enfant = noeud_racine.retrouveEnfant(coup)
                    noeud_racine = noeud_racine.enfants[no_enfant]
                    noeud_racine.devient_racine()
                
                else : 
                    noeud_racine = Node(grille, None, 1, None)
            else :
                coup = robot_joue_encore_mieux(noeud_racine, MCTS_COULEUR)[0]
                print (coup, " coup joué")
                afficher_grille(noeud_racine.etat)
                grille = jouer_coup(couleur.rouge, coup, grille)
                joue = not (verif_gagnage(couleur.rouge, grille))
                if (len(noeud_racine.enfants)>0) : 
                    no_enfant = noeud_racine.retrouveEnfant(coup)
                    noeud_racine = noeud_racine.enfants[no_enfant]
                    noeud_racine.devient_racine()               
                else : 
                    noeud_racine = Node(grille, None, 0, None)
            afficher_grille(grille)
            no_tour = no_tour + 1


def choisirCouleurIAs():
    couleurMinMax = random.choice([couleur.rouge,couleur.jaune])
    if(couleurMinMax==couleur.rouge):
        couleurMCTS=couleur.jaune
        return [couleurMinMax, couleurMCTS]
    else:
        couleurMCTS=couleur.rouge
        return [couleurMinMax, couleurMCTS]


def robot_joue_encore_mieux(noeud_racine, val, iteration): #MCTS
    if (val == couleur.rouge) : 
        val2 = couleur.jaune 
    else :
        val2 = couleur.rouge
    for i in range (iteration) :
        tmp = noeud_racine.selection()
        tmp.extension()
        if (len(tmp.enfants) != 0) :
            tmp = random.choice(tmp.enfants)
            simulation = tmp.simulation()
            tmp.propagationDuResultat(simulation)
        else : 
            
            if (verif_gagnage(val, tmp.etat)):
                simulation = 1
            elif (verif_gagnage(val2, tmp.etat)) :
                simulation = -1
            else : 
                simulation = 0 
            
        tmp.propagationDuResultat(simulation)
    return noeud_racine.meilleurEnfant().mouvement


def jouerMinmaxMCTS(grille, iteration, max_depth) : 
    chemin = os.getcwd()
    print(chemin)
    if os.path.exists( chemin + "\\temps.xlsx") : 
        wb = px.load_workbook(chemin + "\\temps.xlsx")
    else : 
        wb = px.Workbook()
    nom = str(iteration)+ '_' + str(max_depth) + '_' + str(C)
    if not nom in wb.sheetnames :
        wb.create_sheet(nom)
        wb.active = wb[nom]
        
        ws = wb.active
        ws['A1'] = 'Nombre de partie jouée'
        ws['A2'] = 0
        ws['B1'] = 'Numero Tour'
        ws['C1'] = 'nb de passage'
        ws['D1'] = "MCTS temps moyen"
        ws['E1'] = "MINMAX temps moyen" 
        ws['F1'] = "Nombre Victoire MCTS"
        ws['G1'] = "Nombre Victoire Minmax"
        ws['F2'] = 0
        ws['G2'] = 0
    
    else : 
        wb.active = wb[nom]
        
        ws = wb.active
    ws['A2'] = ws['A2'].value + 1
        
    
    joue = True
    no_tour = 0
    noeud_racine = Node(grille, None, 1, None)
    tempsMCTS=0
    tempsMINMAX=0
    compteur = 1
    if(MCTS_COULEUR==couleur.rouge):
        while joue :
            
            
            compteur += no_tour % 2
            ws['B'+str(compteur+2)] = compteur

            if  ws['C'+str(compteur+2)].value == None : 
                ws['C'+str(compteur+2)] = 1
            else : 
                ws['C'+str(compteur+2)] = ws['C'+str(compteur+2)].value + 1 
                
            if (no_tour % 2) == 0 :
                
                tmpMCTS = time.perf_counter()
                coup = robot_joue_encore_mieux(noeud_racine, MCTS_COULEUR, iteration)[0]
                
                tmpMCTS = time.perf_counter() - tmpMCTS
                tempsMCTS += tmpMCTS
                if  ws['D'+str(compteur+2)].value == None : 
                    ws['D'+str(compteur+2)] = tmpMCTS
                else : 
                    ws['D'+str(compteur+2)] = ((tmpMCTS) + (ws['C'+str(compteur+2)].value -1) *  ws['D'+str(compteur+2)].value ) / ws['C'+str(compteur+2)].value
                
                print (coup, " coup joué")
                afficher_grille(noeud_racine.etat)
                grille = jouer_coup(MCTS_COULEUR, coup, grille)
                joue = not (verif_gagnage(MCTS_COULEUR, grille))

                if (len(noeud_racine.enfants)>0) : 
                    no_enfant = noeud_racine.retrouveEnfant(coup)
                    noeud_racine = noeud_racine.enfants[no_enfant]
                    noeud_racine.devient_racine()
                    
                else : 
                    noeud_racine = Node(grille, None, 0, None)
            else :
                tmpMinMax = time.perf_counter()
                coup = robot_joue_mieux(liste_coup_possible(grille), grille, max_depth)
                tmpMinMax = time.perf_counter() - tmpMinMax
                tempsMINMAX += tmpMinMax
                
                if ws['E'+str(compteur+1)].value == None : 
                    ws['E'+str(compteur+1)] = tmpMinMax
                else : 
                    ws['E'+str(compteur+1)] = ((tmpMCTS) + (ws['C'+str(compteur+1)].value - 1) *  ws['E'+str(compteur+1)].value) / ws['C'+str(compteur+1)].value
                
                
                print (coup, " coup joué")
                grille = jouer_coup(MINMAX_COULEUR, coup, grille)
                joue = not (verif_gagnage(MINMAX_COULEUR, grille))
                

                if (len(noeud_racine.enfants)>0) : 
                    no_enfant = noeud_racine.retrouveEnfant(coup)
                    noeud_racine = noeud_racine.enfants[no_enfant]
                    noeud_racine.devient_racine()
                    
                else : 
                    noeud_racine = Node(grille, None, 1, None)
            afficher_grille(grille)
            no_tour = no_tour + 1

            
            
            if grille_est_pleine(grille) : 
                return 0
        
        if (no_tour % 2 == 0) :
            ws['G2'] =  ws['G2'].value + 1
            wb.save(chemin + "\\temps.xlsx") 
            return -1 #victoire de Minmax
        else :
            
            ws['F2'] =  ws['F2'].value + 1
            wb.save(chemin + "\\temps.xlsx") 
            return 1 #victoire de MCTS
    else :  
        while joue :
            
            
            compteur += no_tour % 2
            ws['B'+str(compteur+2)] = compteur

            if  ws['C'+str(compteur+1)].value == None : 
                ws['C'+str(compteur+1)] = 1
            else : 
                ws['C'+str(compteur+1)] = ws['C'+str(compteur+1)].value + 1 
                
            if (no_tour % 2) != 0 :
                
                tmpMCTS = time.perf_counter()
                coup = robot_joue_encore_mieux(noeud_racine, MCTS_COULEUR)[0]
                
                tmpMCTS = time.perf_counter() - tmpMCTS
                tempsMCTS += tmpMCTS
                if  ws['D'+str(compteur+1)].value == None : 
                    ws['D'+str(compteur+1)] = tmpMCTS
                else : 
                    ws['D'+str(compteur+1)] = ((tmpMCTS) + (ws['C'+str(compteur+1)].value -1) *  ws['D'+str(compteur+1)].value ) / ws['C'+str(compteur+1)].value
                
                print (coup, " coup joué")
                afficher_grille(noeud_racine.etat)
                grille = jouer_coup(MCTS_COULEUR, coup, grille)
                joue = not (verif_gagnage(MCTS_COULEUR, grille))

                if (len(noeud_racine.enfants)>0) : 
                    no_enfant = noeud_racine.retrouveEnfant(coup)
                    noeud_racine = noeud_racine.enfants[no_enfant]
                    noeud_racine.devient_racine()
                    
                else : 
                    noeud_racine = Node(grille, None, 0, None)
            else :
                
                
                tmpMinMax = time.perf_counter()
                coup = robot_joue_mieux(liste_coup_possible(grille), grille, max_depth)
                tmpMinMax = time.perf_counter() - tmpMinMax
                tempsMINMAX += tmpMinMax
                
                if ws['E'+str(compteur+2)].value == None : 
                    ws['E'+str(compteur+2)] = tmpMinMax
                else : 
                    if ws['C'+str(compteur+2)].value == None:
                        ws['C'+str(compteur+2)].value = 1
                    
                    ws['E'+str(compteur+2)] = ((tmpMinMax) + (ws['C'+str(compteur+2)].value - 1) *  ws['E'+str(compteur+2)].value) / ws['C'+str(compteur+2)].value
                
                
                print (coup, " coup joué")
                grille = jouer_coup(MINMAX_COULEUR, coup, grille)
                joue = not (verif_gagnage(MINMAX_COULEUR, grille))
                

                if (len(noeud_racine.enfants)>0) : 
                    no_enfant = noeud_racine.retrouveEnfant(coup)
                    noeud_racine = noeud_racine.enfants[no_enfant]
                    noeud_racine.devient_racine()
                    
                else : 
                    noeud_racine = Node(grille, None, 1, None)
            afficher_grille(grille)
            no_tour = no_tour + 1

            
            
            if grille_est_pleine(grille) : 
                return 0
        
        if (no_tour % 2 != 0) :
            ws['G2'] =  ws['G2'].value + 1
            wb.save(chemin + "\\temps.xlsx") 
            return -1 #victoire de Minmax
        else :
            
            ws['F2'] =  ws['F2'].value + 1
            wb.save(chemin + "\\temps.xlsx") 
            return 1 #victoire de MCTS


def humain_joue(grille):
    print("veuiller entre le numero de la colonne (allant de 0 à 6) dans laquelle vous voulez jouer")
    a = saisir(grille)
    #grille = jouer_coup(couleur.rouge, a, grille)
    return a

def saisir(grille):
    a = input()
    if not (a.isnumeric()):
        a = saisir(grille)
    a = int(a)
    if a < 0 or a > 6:
        print(
            "Cette colonne n'existe pas, rentrez a nouveau une colonne (allant de 0 à 6) dans laquelle vous voulez jouer")
        a = saisir(grille)
    if grille[a][5] != ' ':
        a = saisir(grille)
    return a


def choixModeDeJeu():
    print("Veuillez choisir le mode de jeu")
    print("Tapez 1 pour jouer contre MinMax, tapez 2 pour jouer contre MCTS, et tapez 3 pour se faire affronter les deux IAs")
    choix=input()
    if(not(choix.isnumeric())):
        ("erreur, entrez un nombre")
        choixModeDeJeu()
    print(choix)
    choix = int(choix)
    if(choix!=1 and choix!=2 and choix!=3):
        print("erreur, veuillez réessayez")
        choixModeDeJeu()
    return choix


def jouer() :
    choix=choixModeDeJeu()
    if(choix==1):
        print("HUMAIN VS MinMax")
        print("Choisissez la profondeur de MinMax (pas plus de 5)")
        MAX_DEPTH_MINMAX=input()
        if(not(MAX_DEPTH_MINMAX.isnumeric())):
            print("Vous n'avez pas entré de nombre. Choisissez la profondeur de MinMax (pas plus de 5)")
            MAX_DEPTH_MINMAX=input() 
        MAX_DEPTH_MINMAX = int(MAX_DEPTH_MINMAX)         
        if(MAX_DEPTH_MINMAX!=1 and MAX_DEPTH_MINMAX!=2 and MAX_DEPTH_MINMAX!=3 and MAX_DEPTH_MINMAX!=4 and MAX_DEPTH_MINMAX!=5):
            print("Choisissez la profondeur de MinMax (pas plus de 5)")
            MAX_DEPTH_MINMAX=input()               
        listeCouleur=choisirCouleurIAs()
        MCTS_COULEUR = listeCouleur[0]
        MINMAX_COULEUR = listeCouleur[1]
        if(MINMAX_COULEUR==couleur.rouge):
            print("MINMAX Commence")
        else:
            print("Vous commencez")
        grille_jeu=grille_initialisation()
        jouerMinMax(grille_jeu)

    elif(choix==2):
        print("HUMAIN VS MCTS")
        print("Choisissez le nombre d'itérations de MCTS (cela doit-être un entier strictement positif")
        MCTS_ITERATION=input()
        MCTS_ITERATION = int(MCTS_ITERATION)
        #listeCouleur=choisirCouleurIAs()
        #MCTS_COULEUR = listeCouleur[0]
        #MINMAX_COULEUR = listeCouleur[1]
        MCTS_COULEUR=couleur.jaune # pour une raison que j'ignore, quand on appelle jouerMCTS, MCTS joue n'importe quand il débute
        if(MCTS_COULEUR==couleur.rouge):
            print("MCTS Commence")
        else:
            print("Vous commencez")
        grille_jeu=grille_initialisation()
        jouerMCTS(grille_jeu, MCTS_ITERATION)

    elif(choix==3):
        print("MCTS VS MinMax")
        print("Choisissez la profondeur de MinMax (pas plus de 5)")
        MAX_DEPTH_MINMAX=input()

        if(not(MAX_DEPTH_MINMAX.isnumeric())):
            print("Choisissez la profondeur de MinMax (pas plus de 5)")
            MAX_DEPTH_MINMAX=input() 

        MAX_DEPTH_MINMAX = int(MAX_DEPTH_MINMAX)
        if(MAX_DEPTH_MINMAX!=1 and MAX_DEPTH_MINMAX!=2 and MAX_DEPTH_MINMAX!=3 and MAX_DEPTH_MINMAX!=4 and MAX_DEPTH_MINMAX!=5):
                print("Choisissez la profondeur de MinMax (pas plus de 5)")
                MAX_DEPTH_MINMAX=input()
        print("Choisissez le nombre d'itérations de MCTS (cela doit-être un entier strictement positif")
        MCTS_ITERATION=input()

        if(not(MCTS_ITERATION.isnumeric() and int(MCTS_ITERATION)>0)):
            print("Choisissez le nombre d'itérations de MCTS (cela doit-être un entier strictement positif")
            MCTS_ITERATION=input()
        MCTS_ITERATION = int(MCTS_ITERATION)
        print("Choississez le nombre de parties à effectuer (entier strictement positif")
        nbrParties=input()

        if(not(nbrParties.isnumeric() and int(nbrParties)>0)):
            print("Choississez le nombre de parties à effectuer (entier strictement positif")
            nbrParties=input() 
            
        nbrParties=int(nbrParties)
        print("le résumé des résultats sera retranscrit dans un fichier qui sera dans votre répertoire") 
        try : 
            os.mkdir("res")
        except : 
            pass                         
        fichier = open("res/result_"+str(MAX_DEPTH_MINMAX)+"_"+str(MCTS_ITERATION)+"_.txt", "a")
        fichier.write("L'IA avec les x est l'IA qui joue en première " + "\n") 
        listeScore = []


        for i in range (nbrParties) :
            listeCouleur=choisirCouleurIAs()
            MCTS_COULEUR = listeCouleur[0]
            MINMAX_COULEUR = listeCouleur[1]
            if(MINMAX_COULEUR==couleur.rouge):
                print("MINMAX Commence")
            else:
                print("MCTS Commence")
            fichier.write("MCTS: " + MCTS_COULEUR.value + "\nMinMax: " + MINMAX_COULEUR.value + "\n") 
            grille_jeu = grille_initialisation()
            res = jouerMinmaxMCTS(grille_jeu, MCTS_ITERATION, MAX_DEPTH_MINMAX)
            listeScore.append(res)
            fichier.write(returnString(grille_jeu))
            fichier.write("Pour MCTS de " + str(MCTS_ITERATION) + " d'iteration et MinMax a " + str(MAX_DEPTH_MINMAX) + " de profondeur : \n")
            fichier.write("Nombre de victoire de MinMax : " + str(listeScore.count(-1)) + "\n")
            fichier.write("Nombre de victoire de MCTS : " + str(listeScore.count(1)) + "\n")
            moy = sum(listeScore) / len(listeScore)